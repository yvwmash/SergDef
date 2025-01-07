import numpy   as np

from datetime                 import date
from openpyxl                 import Workbook as xl_wb
from openpyxl                 import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles          import NamedStyle as xl_style, Font as xl_font, Border as xl_border, Side as xl_side 
from openpyxl.styles          import PatternFill as xl_fill, GradientFill as xl_gradfill
from openpyxl.styles          import Alignment as xl_align

from h_types     import *

# ###
def autosize_ws(ws):
 for col in ws.columns:
  max_len = 0
  column = col[0].column_letter # column name
  for cell in col:
   if len(str(cell.value)) > max_len:
    max_len = len(str(cell.value))
  adjusted_width = (max_len + 2) * 1.2
  ws.column_dimensions[column].width = adjusted_width

# ###
def save_e(fn, d_nm2df, need_date, need_index, need_autosize = true, c_map = none):
 wb = xl_wb()
 wsa = wb.active
 wb.remove(wsa)
 
 stoday = ''
 if need_date:
  stoday = date.today().strftime('%d.%m.%Y') + '_'
  fn = fn + stoday

 if c_map:
  for df_nm, m in c_map:
   df = d_nm2df[df_nm]
   for c_nm, c_typ in c_map[df_nm]:
    df[c_nm] = c_conv(df[c_nm], c_typ)

 for ws_nm, df in d_nm2df.items(): # loop through `dict` of dataframes
  ws = wb.create_sheet()
  ws.title = ws_nm
  df = df.fillna(np.nan)
  for r in dataframe_to_rows(df, index = need_index, header = true):
   ws.append(r)
  autosize_ws(ws)
 fn = '/mnt/hdd/db_io/from_scripts/' + fn + '.xlsx'
 wb.save(fn)
 return fn
